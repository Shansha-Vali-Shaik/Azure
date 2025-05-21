import pandas as pd
from datetime import timedelta
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
 
class InFloorReportGenerator:
    def __init__(self, excel_path, input_sheet="Raw", output_sheet="InFloor_Summary"):
        self.excel_path = excel_path
        self.input_sheet = input_sheet
        self.output_sheet = output_sheet
        self.df = None
        self.summary_df = None
 
    def load_and_preprocess(self):
        # Read Excel input sheet
        self.df = pd.read_excel(self.excel_path, sheet_name=self.input_sheet, engine='openpyxl')
 
        # Preprocess
        self.df["Date"] = pd.to_datetime(self.df["Date"], format="%d-%m-%Y %H:%M")
        self.df["Action"] = self.df["Device"].str.upper().apply(
            lambda x: "ENTRANCE" if "ENTRANCE" in x else ("EXIT" if "EXIT" in x else None)
        )
        self.df = self.df[self.df["Action"].isin(["ENTRANCE", "EXIT"])].copy()
        self.df = self.df.sort_values("Date").reset_index(drop=True)
 
    def generate_summary(self):
        results = []
 
        for user, user_df in self.df.groupby("User"):
            user_df = user_df.sort_values("Date").reset_index(drop=True)
            user_df["Day"] = user_df["Date"].dt.date
 
            sessions = []
            i = 0
            while i < len(user_df):
                row = user_df.iloc[i]
                if row["Action"] == "ENTRANCE":
                    in_time = row["Date"]
                    j = i + 1
                    while j < len(user_df):
                        next_row = user_df.iloc[j]
                        if next_row["Action"] == "EXIT":
                            out_time = next_row["Date"]
                            sessions.append((in_time.date(), in_time, out_time))
                            break
                        j += 1
                    i = j
                else:
                    i += 1
 
            session_df = pd.DataFrame(sessions, columns=["Day", "In", "Out"])
 
            if session_df.empty:
                continue
 
            grouped = session_df.groupby("Day")
            for day, grp in grouped:
                login_time = grp["In"].min()
                logout_time = grp["Out"].max()
                duration = logout_time - login_time
                in_floor_time = (grp["Out"] - grp["In"]).sum()
                break_time = duration - in_floor_time
 
                results.append({
                    "User": user,
                    "Date": day,
                    "Login Time": login_time.time(),
                    "Logout Time": logout_time.time(),
                    "Total Duration": duration,
                    "In-Floor Time": in_floor_time,
                    "Break Time": break_time
                })
 
        self.summary_df = pd.DataFrame(results)
        self.summary_df = self.summary_df.sort_values(["User", "Date"]).reset_index(drop=True)
 
    def export_to_excel(self):
        if self.summary_df is None or self.summary_df.empty:
            print("No data to export.")
            return
 
        try:
            wb = load_workbook(self.excel_path)
        except FileNotFoundError:
            wb = Workbook()
            del wb[wb.sheetnames[0]]
 
        # Delete output sheet if exists
        if self.output_sheet in wb.sheetnames:
            del wb[self.output_sheet]
 
        ws = wb.create_sheet(self.output_sheet)
 
        for r_idx, row in enumerate(dataframe_to_rows(self.summary_df, index=False, header=True), 1):
            ws.append(row)
            if r_idx == 1:
                continue  # header row
 
            ws.cell(row=r_idx, column=3).number_format = 'hh:mm:ss'      # Login Time
            ws.cell(row=r_idx, column=4).number_format = 'hh:mm:ss'      # Logout Time
            ws.cell(row=r_idx, column=5).number_format = '[h]:mm:ss'     # Total Duration
            ws.cell(row=r_idx, column=6).number_format = '[h]:mm:ss'     # In-Floor Time
            ws.cell(row=r_idx, column=7).number_format = '[h]:mm:ss'     # Break Time
 
        wb.save(self.excel_path)
 
    def run(self):
        self.load_and_preprocess()
        self.generate_summary()
        self.export_to_excel()
        print(self.summary_df.head(10))
 
# Usage example:
generator = InFloorReportGenerator(r"C:\Users\shansha.vali\Documents\TL_Break_Report\Report\Report_File.xlsx", input_sheet="Raw")
generator.run()
